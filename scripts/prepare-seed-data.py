#!/usr/bin/env python3.11
"""
Convert CMC_Go_Sample_Data_Regenerated.xlsx to JSON seed data
"""
import json
import openpyxl
from datetime import datetime

# Load the Excel file
wb = openpyxl.load_workbook('/home/ubuntu/upload/CMC_Go_Sample_Data_Regenerated.xlsx')

# Process Districts
districts_sheet = wb['RegionsDistricts']
districts = []
seen_slugs = set()

for row in districts_sheet.iter_rows(min_row=2, values_only=True):
    region, district_name, district_slug = row
    if district_slug and district_slug not in seen_slugs:
        districts.append({
            'id': district_slug,
            'name': district_name,
            'region': region
        })
        seen_slugs.add(district_slug)

# Process Campuses
campuses_sheet = wb['Campuses']
campuses = []

for row in campuses_sheet.iter_rows(min_row=2, values_only=True):
    campus_id, campus_name, district, district_slug, region = row
    if campus_id:
        campuses.append({
            'id': int(campus_id),
            'name': campus_name,
            'districtId': district_slug
        })

# Process People
people_sheet = wb['People']
people = []
needs = []
notes = []

person_id_counter = 1
need_id_counter = 1
note_id_counter = 1

for row in people_sheet.iter_rows(min_row=2, values_only=True):
    person_id, full_name, role, region, district, district_slug, campus, status, need_type, need_amount, note_text, last_updated = row
    
    if not full_name:
        continue
    
    # Find campus ID by matching campus name and district
    campus_id = None
    for c in campuses:
        if c['name'] == campus and c['districtId'] == district_slug:
            campus_id = c['id']
            break
    
    if not campus_id:
        continue
    
    # Convert last_updated to ISO format
    if isinstance(last_updated, datetime):
        last_updated_str = last_updated.strftime('%Y-%m-%d %H:%M:%S')
    else:
        last_updated_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Add person
    people.append({
        'id': person_id_counter,
        'name': full_name,
        'campusId': campus_id,
        'districtId': district_slug,
        'status': status if status else 'Not invited yet',
        'role': role if role else None,
        'lastUpdated': last_updated_str
    })
    
    # Add need if exists
    if need_type:
        need_data = {
            'id': need_id_counter,
            'personId': person_id_counter,
            'type': need_type,
            'notes': None,
            'isActive': True
        }
        if need_type == 'Financial' and need_amount:
            # Convert to cents
            need_data['amount'] = int(float(need_amount) * 100)
        needs.append(need_data)
        need_id_counter += 1
    
    # Add note if exists
    if note_text:
        notes.append({
            'id': note_id_counter,
            'personId': person_id_counter,
            'text': note_text,
            'isLeaderOnly': False
        })
        note_id_counter += 1
    
    person_id_counter += 1

# Write to JSON files
with open('/home/ubuntu/cmc-go/scripts/seed-districts.json', 'w') as f:
    json.dump(districts, f, indent=2)

with open('/home/ubuntu/cmc-go/scripts/seed-campuses.json', 'w') as f:
    json.dump(campuses, f, indent=2)

with open('/home/ubuntu/cmc-go/scripts/seed-people.json', 'w') as f:
    json.dump(people, f, indent=2)

with open('/home/ubuntu/cmc-go/scripts/seed-needs.json', 'w') as f:
    json.dump(needs, f, indent=2)

with open('/home/ubuntu/cmc-go/scripts/seed-notes.json', 'w') as f:
    json.dump(notes, f, indent=2)

print(f"Generated seed data:")
print(f"  {len(districts)} districts")
print(f"  {len(campuses)} campuses")
print(f"  {len(people)} people")
print(f"  {len(needs)} needs")
print(f"  {len(notes)} notes")
